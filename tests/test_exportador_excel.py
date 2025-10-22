import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.exportador_excel import ExportadorExcel, para_excel_formatado


class TestExcelExporter:

    @pytest.fixture
    def exemplo_dataframe(self):
        """Cria um DataFrame de exemplo com diferentes tipos de dados."""
        return pd.DataFrame(
            {
                "Nome": ["João", "Maria", "Pedro"],
                "Salário": [1500.50, 2300.75, 1800.00],
                "Data": pd.to_datetime(["2024-01-01", "2024-02-01", "2024-03-01"]),
                "Idade": [25, 30, 28],
                "Ativo": [True, False, True],
                "Observações": ["Texto", None, "Outro texto"],
            }
        )

    @pytest.fixture
    def arquivo_excel_temporario(self):
        """Cria um arquivo Excel temporário para testes."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        yield temp_file.name
        # Cleanup
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

    def test_inicializacao_exportador(self):
        """Testa se o exportador é inicializado corretamente."""
        exporter = ExportadorExcel()
        assert exporter.estilo_numerico.name == "formato_numerico"
        assert exporter.estilo_numerico.number_format == "#,##0.00"

    def test_deteccao_valor_float(self):
        """Testa a detecção de valores float."""
        exporter = ExportadorExcel()

        # Valores que devem ser detectados como float
        assert exporter._eh_float(123.45) == True
        assert exporter._eh_float(-50.5) == True
        assert exporter._eh_float(0.0) == True
        assert exporter._eh_float(1500.50) == True

        # Valores que NÃO devem ser detectados como float
        assert exporter._eh_float(123) == False  # int
        assert exporter._eh_float(0) == False  # int
        assert exporter._eh_float("123.45") == False  # string
        assert exporter._eh_float("texto") == False
        assert exporter._eh_float(None) == False
        assert exporter._eh_float(pd.NA) == False
        assert exporter._eh_float(True) == False  # bool

    def test_funcionalidade_basica_to_excel(
        self, exemplo_dataframe, arquivo_excel_temporario
    ):
        """Testa a funcionalidade básica de exportação."""
        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(exemplo_dataframe, writer, "Teste", index=False)

        # Verifica se o arquivo foi criado
        assert os.path.exists(arquivo_excel_temporario)

        # Carrega o arquivo e verifica se os dados estão corretos
        df_carregado = pd.read_excel(arquivo_excel_temporario, sheet_name="Teste")
        assert len(df_carregado) == 3
        assert list(df_carregado.columns) == list(exemplo_dataframe.columns)

    def test_formatacao_apenas_floats(self, arquivo_excel_temporario):
        """Testa se a formatação é aplicada apenas em floats."""
        df = pd.DataFrame(
            {
                "Texto": ["A", "B", "C"],
                "Float": [1234.56, 7890.12, 5000.00],
                "Inteiro": [100, 200, 300],
                "Boolean": [True, False, True],
            }
        )

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df, writer, "Formatação", index=False)

        # Carrega o workbook diretamente para verificar a formatação
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["Formatação"]

        # Verifica se o estilo de número foi adicionado
        assert "formato_numerico" in workbook.named_styles

        # Apenas as células float devem ter formatação
        # Células B2, B3, B4 (coluna Float)
        assert worksheet["B2"].style == "formato_numerico"
        assert worksheet["B3"].style == "formato_numerico"
        assert worksheet["B4"].style == "formato_numerico"

        # Células de outros tipos não devem ter formatação especial
        assert worksheet["A2"].style != "formato_numerico"  # Texto
        assert worksheet["A3"].style != "formato_numerico"
        assert worksheet["A4"].style != "formato_numerico"

        assert worksheet["C2"].style != "formato_numerico"  # Inteiro
        assert worksheet["C3"].style != "formato_numerico"
        assert worksheet["C4"].style != "formato_numerico"

        assert worksheet["D2"].style != "formato_numerico"  # Boolean
        assert worksheet["D3"].style != "formato_numerico"
        assert worksheet["D4"].style != "formato_numerico"

    def test_to_excel_com_indice(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa exportação com índice incluído."""
        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(exemplo_dataframe, writer, "ComIndice", index=True)

        # Carrega e verifica
        df_carregado = pd.read_excel(
            arquivo_excel_temporario, sheet_name="ComIndice", index_col=0
        )
        assert len(df_carregado) == 3
        assert len(df_carregado.columns) == len(exemplo_dataframe.columns)

    def test_funcao_conveniencia(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa a função de conveniência to_excel_formatted."""
        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            para_excel_formatado(exemplo_dataframe, writer, "Conveniência", index=False)

        # Verifica se funcionou corretamente
        assert os.path.exists(arquivo_excel_temporario)
        df_carregado = pd.read_excel(
            arquivo_excel_temporario, sheet_name="Conveniência"
        )
        assert len(df_carregado) == 3

    def test_multiplas_planilhas(self, arquivo_excel_temporario):
        """Testa exportação para múltiplas planilhas."""
        df1 = pd.DataFrame({"A": [1, 2], "B": [10.5, 20.5]})
        df2 = pd.DataFrame({"C": [3, 4], "D": [30.5, 40.5]})

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df1, writer, "Planilha1", index=False)
            exporter.para_excel(df2, writer, "Planilha2", index=False)

        # Verifica se ambas as planilhas existem
        workbook = load_workbook(arquivo_excel_temporario)
        assert "Planilha1" in workbook.sheetnames
        assert "Planilha2" in workbook.sheetnames

        # Verifica formatação em ambas
        ws1 = workbook["Planilha1"]
        ws2 = workbook["Planilha2"]

        # Apenas floats devem ter formatação
        assert ws1["A2"].style != "formato_numerico"  # int
        assert ws1["B2"].style == "formato_numerico"  # float
        assert ws2["A3"].style != "formato_numerico"  # int
        assert ws2["B3"].style == "formato_numerico"  # float

    def test_dataframe_vazio(self, arquivo_excel_temporario):
        """Testa comportamento com DataFrame vazio."""
        df_vazio = pd.DataFrame()
        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df_vazio, writer, "Vazio", index=False)

        # Deve funcionar sem erros
        assert os.path.exists(arquivo_excel_temporario)

        # Verifica se a planilha foi criada
        workbook = load_workbook(arquivo_excel_temporario)
        assert "Vazio" in workbook.sheetnames

    def test_dataframe_com_valores_nan(self, arquivo_excel_temporario):
        """Testa comportamento com valores NaN."""
        df_nan = pd.DataFrame({"Floats": [1.5, None, 3.5], "Texto": ["A", "B", None]})

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df_nan, writer, "ComNaN", index=False)

        # Carrega o workbook para verificar formatação
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["ComNaN"]

        # Apenas valores float válidos devem ter formatação
        assert worksheet["A2"].style == "formato_numerico"  # 1.5
        assert worksheet["A3"].style != "formato_numerico"  # None
        assert worksheet["A4"].style == "formato_numerico"  # 3.5

        # Células de texto não devem ter formatação
        assert worksheet["B2"].style != "formato_numerico"  # 'A'
        assert worksheet["B3"].style != "formato_numerico"  # 'B'
        assert worksheet["B4"].style != "formato_numerico"  # None

    def test_kwargs_adicionais(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa se kwargs adicionais são passados corretamente."""
        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(
                exemplo_dataframe,
                writer,
                "ComKwargs",
                index=False,
                startrow=2,
                startcol=1,
            )

        # Verifica se funcionou (dados começam na linha 3, coluna B)
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["ComKwargs"]

        # O cabeçalho deve estar na linha 3 (startrow=2, 0-indexed)
        assert worksheet["B3"].value == "Nome"
        assert worksheet["C3"].value == "Salário"

    def test_apenas_floats_formatados_exemplo_real(self, arquivo_excel_temporario):
        """Testa que apenas floats são formatados em um exemplo mais realista."""
        df = pd.DataFrame(
            {
                "CM": [1001, 1002, 1003],  # int - não formatar
                "Nome": ["João", "Maria", "Pedro"],  # string - não formatar
                "Salário": [1500.50, 2300.75, 1800.00],  # float - formatar
                "Idade": [25, 30, 28],  # int - não formatar
                "Ativo": [True, False, True],  # bool - não formatar
                "Bonus": [150.25, 230.50, 180.75],  # float - formatar
            }
        )

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df, writer, "Exemplo", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["Exemplo"]

        # Apenas as colunas de float devem ter formatação
        for row in range(2, 5):  # Linhas 2-4 (dados)
            assert worksheet[f"A{row}"].style != "formato_numerico"  # CM (int)
            assert worksheet[f"B{row}"].style != "formato_numerico"  # Nome (string)
            assert worksheet[f"C{row}"].style == "formato_numerico"  # Salário (float)
            assert worksheet[f"D{row}"].style != "formato_numerico"  # Idade (int)
            assert worksheet[f"E{row}"].style != "formato_numerico"  # Ativo (bool)
            assert worksheet[f"F{row}"].style == "formato_numerico"  # Bonus (float)

    def test_wrap_text_cabecalhos(self, arquivo_excel_temporario):
        """Testa se wrap text é aplicado nos cabeçalhos."""
        df = pd.DataFrame(
            {
                "Nome Completo do Funcionário": ["João", "Maria"],
                "Salário Base Mensal": [1500.50, 2300.75],
                "Data de Admissão": ["2020-01-01", "2019-05-15"],
            }
        )

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df, writer, "WrapText", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["WrapText"]

        # Verifica se wrap_text está ativado nos cabeçalhos
        assert worksheet["A1"].alignment.wrap_text == True
        assert worksheet["B1"].alignment.wrap_text == True
        assert worksheet["C1"].alignment.wrap_text == True

        # Verifica alinhamento centralizado
        assert worksheet["A1"].alignment.horizontal == "center"
        assert worksheet["A1"].alignment.vertical == "center"

    def test_auto_fit_colunas(self, arquivo_excel_temporario):
        """Testa se as colunas são ajustadas automaticamente."""
        df = pd.DataFrame(
            {
                "A": ["Pequeno", "Texto muito muito longo para testar auto-fit"],
                "B": ["X", "Y"],
                "C": [123.45, 6789.12],
            }
        )

        exporter = ExportadorExcel()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.para_excel(df, writer, "AutoFit", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["AutoFit"]

        # Verifica se as larguras foram definidas
        col_a_width = worksheet.column_dimensions["A"].width
        col_b_width = worksheet.column_dimensions["B"].width
        col_c_width = worksheet.column_dimensions["C"].width

        # Coluna A deve ser a mais larga (texto longo)
        assert col_a_width > col_b_width
        assert col_a_width > col_c_width
