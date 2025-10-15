import os
import tempfile

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.excel_exporter import ExcelExporter, to_excel_formatted


class TestExcelExporter:

    @pytest.fixture
    def exemplo_dataframe(self):
        """Cria um DataFrame de exemplo com diferentes tipos de dados."""
        return pd.DataFrame(
            {
                "Nome": ["João", "Maria", "Pedro"],
                "Salário": [1500.50, 2300.75, 1800.00],
                "Data": pd.to_datetime(["2024-01-01", "2024-02-01", "2024-03-01"]),
                "Idade": [25, 30, 28],
                "Ativo": [True, False, True],
                "Observações": ["Texto", None, "Outro texto"],
            }
        )

    @pytest.fixture
    def arquivo_excel_temporario(self):
        """Cria um arquivo Excel temporário para testes."""
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        temp_file.close()
        yield temp_file.name
        # Cleanup
        if os.path.exists(temp_file.name):
            os.unlink(temp_file.name)

    def test_inicializacao_exportador(self):
        """Testa se o exportador é inicializado corretamente."""
        exporter = ExcelExporter()
        assert exporter.number_style.name == "number_format"
        assert exporter.number_style.number_format == "#,##0.00"

    def test_deteccao_valor_float(self):
        """Testa a detecção de valores float."""
        exporter = ExcelExporter()

        # Valores que devem ser detectados como float
        assert exporter._is_float_value(123.45) == True
        assert exporter._is_float_value(-50.5) == True
        assert exporter._is_float_value(0.0) == True
        assert exporter._is_float_value(1500.50) == True

        # Valores que NÃO devem ser detectados como float
        assert exporter._is_float_value(123) == False  # int
        assert exporter._is_float_value(0) == False  # int
        assert exporter._is_float_value("123.45") == False  # string
        assert exporter._is_float_value("texto") == False
        assert exporter._is_float_value(None) == False
        assert exporter._is_float_value(pd.NA) == False
        assert exporter._is_float_value(True) == False  # bool

    def test_funcionalidade_basica_to_excel(
        self, exemplo_dataframe, arquivo_excel_temporario
    ):
        """Testa a funcionalidade básica de exportação."""
        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(exemplo_dataframe, writer, "Teste", index=False)

        # Verifica se o arquivo foi criado
        assert os.path.exists(arquivo_excel_temporario)

        # Carrega o arquivo e verifica se os dados estão corretos
        df_carregado = pd.read_excel(arquivo_excel_temporario, sheet_name="Teste")
        assert len(df_carregado) == 3
        assert list(df_carregado.columns) == list(exemplo_dataframe.columns)

    def test_formatacao_apenas_floats(self, arquivo_excel_temporario):
        """Testa se a formatação é aplicada apenas em floats."""
        df = pd.DataFrame(
            {
                "Texto": ["A", "B", "C"],
                "Float": [1234.56, 7890.12, 5000.00],
                "Inteiro": [100, 200, 300],
                "Boolean": [True, False, True],
            }
        )

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df, writer, "Formatação", index=False)

        # Carrega o workbook diretamente para verificar a formatação
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["Formatação"]

        # Verifica se o estilo de número foi adicionado
        assert "number_format" in workbook.named_styles

        # Apenas as células float devem ter formatação
        # Células B2, B3, B4 (coluna Float)
        assert worksheet["B2"].style == "number_format"
        assert worksheet["B3"].style == "number_format"
        assert worksheet["B4"].style == "number_format"

        # Células de outros tipos não devem ter formatação especial
        assert worksheet["A2"].style != "number_format"  # Texto
        assert worksheet["A3"].style != "number_format"
        assert worksheet["A4"].style != "number_format"

        assert worksheet["C2"].style != "number_format"  # Inteiro
        assert worksheet["C3"].style != "number_format"
        assert worksheet["C4"].style != "number_format"

        assert worksheet["D2"].style != "number_format"  # Boolean
        assert worksheet["D3"].style != "number_format"
        assert worksheet["D4"].style != "number_format"

    def test_to_excel_com_indice(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa exportação com índice incluído."""
        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(exemplo_dataframe, writer, "ComIndice", index=True)

        # Carrega e verifica
        df_carregado = pd.read_excel(
            arquivo_excel_temporario, sheet_name="ComIndice", index_col=0
        )
        assert len(df_carregado) == 3
        assert len(df_carregado.columns) == len(exemplo_dataframe.columns)

    def test_funcao_conveniencia(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa a função de conveniência to_excel_formatted."""
        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            to_excel_formatted(exemplo_dataframe, writer, "Conveniência", index=False)

        # Verifica se funcionou corretamente
        assert os.path.exists(arquivo_excel_temporario)
        df_carregado = pd.read_excel(
            arquivo_excel_temporario, sheet_name="Conveniência"
        )
        assert len(df_carregado) == 3

    def test_multiplas_planilhas(self, arquivo_excel_temporario):
        """Testa exportação para múltiplas planilhas."""
        df1 = pd.DataFrame({"A": [1, 2], "B": [10.5, 20.5]})
        df2 = pd.DataFrame({"C": [3, 4], "D": [30.5, 40.5]})

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df1, writer, "Planilha1", index=False)
            exporter.to_excel(df2, writer, "Planilha2", index=False)

        # Verifica se ambas as planilhas existem
        workbook = load_workbook(arquivo_excel_temporario)
        assert "Planilha1" in workbook.sheetnames
        assert "Planilha2" in workbook.sheetnames

        # Verifica formatação em ambas
        ws1 = workbook["Planilha1"]
        ws2 = workbook["Planilha2"]

        # Apenas floats devem ter formatação
        assert ws1["A2"].style != "number_format"  # int
        assert ws1["B2"].style == "number_format"  # float
        assert ws2["A3"].style != "number_format"  # int
        assert ws2["B3"].style == "number_format"  # float

    def test_dataframe_vazio(self, arquivo_excel_temporario):
        """Testa comportamento com DataFrame vazio."""
        df_vazio = pd.DataFrame()
        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df_vazio, writer, "Vazio", index=False)

        # Deve funcionar sem erros
        assert os.path.exists(arquivo_excel_temporario)

        # Verifica se a planilha foi criada
        workbook = load_workbook(arquivo_excel_temporario)
        assert "Vazio" in workbook.sheetnames

    def test_dataframe_com_valores_nan(self, arquivo_excel_temporario):
        """Testa comportamento com valores NaN."""
        df_nan = pd.DataFrame({"Floats": [1.5, None, 3.5], "Texto": ["A", "B", None]})

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df_nan, writer, "ComNaN", index=False)

        # Carrega o workbook para verificar formatação
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["ComNaN"]

        # Apenas valores float válidos devem ter formatação
        assert worksheet["A2"].style == "number_format"  # 1.5
        assert worksheet["A3"].style != "number_format"  # None
        assert worksheet["A4"].style == "number_format"  # 3.5

        # Células de texto não devem ter formatação
        assert worksheet["B2"].style != "number_format"  # 'A'
        assert worksheet["B3"].style != "number_format"  # 'B'
        assert worksheet["B4"].style != "number_format"  # None

    def test_kwargs_adicionais(self, exemplo_dataframe, arquivo_excel_temporario):
        """Testa se kwargs adicionais são passados corretamente."""
        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(
                exemplo_dataframe,
                writer,
                "ComKwargs",
                index=False,
                startrow=2,
                startcol=1,
            )

        # Verifica se funcionou (dados começam na linha 3, coluna B)
        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["ComKwargs"]

        # O cabeçalho deve estar na linha 3 (startrow=2, 0-indexed)
        assert worksheet["B3"].value == "Nome"
        assert worksheet["C3"].value == "Salário"

    def test_apenas_floats_formatados_exemplo_real(self, arquivo_excel_temporario):
        """Testa que apenas floats são formatados em um exemplo mais realista."""
        df = pd.DataFrame(
            {
                "CM": [1001, 1002, 1003],  # int - não formatar
                "Nome": ["João", "Maria", "Pedro"],  # string - não formatar
                "Salário": [1500.50, 2300.75, 1800.00],  # float - formatar
                "Idade": [25, 30, 28],  # int - não formatar
                "Ativo": [True, False, True],  # bool - não formatar
                "Bonus": [150.25, 230.50, 180.75],  # float - formatar
            }
        )

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df, writer, "Exemplo", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["Exemplo"]

        # Apenas as colunas de float devem ter formatação
        for row in range(2, 5):  # Linhas 2-4 (dados)
            assert worksheet[f"A{row}"].style != "number_format"  # CM (int)
            assert worksheet[f"B{row}"].style != "number_format"  # Nome (string)
            assert worksheet[f"C{row}"].style == "number_format"  # Salário (float)
            assert worksheet[f"D{row}"].style != "number_format"  # Idade (int)
            assert worksheet[f"E{row}"].style != "number_format"  # Ativo (bool)
            assert worksheet[f"F{row}"].style == "number_format"  # Bonus (float)

    def test_wrap_text_cabecalhos(self, arquivo_excel_temporario):
        """Testa se wrap text é aplicado nos cabeçalhos."""
        df = pd.DataFrame(
            {
                "Nome Completo do Funcionário": ["João", "Maria"],
                "Salário Base Mensal": [1500.50, 2300.75],
                "Data de Admissão": ["2020-01-01", "2019-05-15"],
            }
        )

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df, writer, "WrapText", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["WrapText"]

        # Verifica se wrap_text está ativado nos cabeçalhos
        assert worksheet["A1"].alignment.wrap_text == True
        assert worksheet["B1"].alignment.wrap_text == True
        assert worksheet["C1"].alignment.wrap_text == True

        # Verifica alinhamento centralizado
        assert worksheet["A1"].alignment.horizontal == "center"
        assert worksheet["A1"].alignment.vertical == "center"

    def test_auto_fit_colunas(self, arquivo_excel_temporario):
        """Testa se as colunas são ajustadas automaticamente."""
        df = pd.DataFrame(
            {
                "A": ["Pequeno", "Texto muito muito longo para testar auto-fit"],
                "B": ["X", "Y"],
                "C": [123.45, 6789.12],
            }
        )

        exporter = ExcelExporter()

        with pd.ExcelWriter(arquivo_excel_temporario, engine="openpyxl") as writer:
            exporter.to_excel(df, writer, "AutoFit", index=False)

        workbook = load_workbook(arquivo_excel_temporario)
        worksheet = workbook["AutoFit"]

        # Verifica se as larguras foram definidas
        col_a_width = worksheet.column_dimensions["A"].width
        col_b_width = worksheet.column_dimensions["B"].width
        col_c_width = worksheet.column_dimensions["C"].width

        # Coluna A deve ser a mais larga (texto longo)
        assert col_a_width > col_b_width
        assert col_a_width > col_c_width
