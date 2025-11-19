import os
from dataclasses import dataclass
from datetime import date
from types import SimpleNamespace

import openpyxl
import pandas as pd

from reenquadramento.avaliacoes import Avaliacao, Avaliacoes
from reenquadramento.dados_faltantes_aeros import (
    DadosFaltantesAeros,
    obtem_dados_faltantes_aeros,
)
from src.carreira import Progressao
from src.classe import Classe
from src.cmbh import CMBH
from src.funcionario import Funcionario
from src.importador_excel import ImportadorProjecaoExcel
from src.nivel import Nivel


@dataclass
class ProgressaoSimulada:
    data: date
    nivel: Nivel
    especial_concedida: bool
    nota_media_especial: float
    nota_avaliacao: float


class TrajetoriaSimulada:
    def __init__(
        self,
        funcionario: Funcionario,
        data_migracao: date,
        progressoes: list[ProgressaoSimulada],
        dados_faltantes_aeros: DadosFaltantesAeros,
    ):
        self.funcionario: Funcionario = funcionario
        self.data_migracao: date = data_migracao
        self.progressoes: list[ProgressaoSimulada] = progressoes
        self.dados_faltantes_aeros: DadosFaltantesAeros = dados_faltantes_aeros

    def escreve(self, writer: pd.ExcelWriter) -> None:
        """Escreve esta trajetória simulada em uma planilha do Excel usando o
        ExcelWriter passado como `writer`. O nome da planilha será o CM do
        funcionário (string).
        """
        workbook = writer.book
        sheet_name = str(self.funcionario.cm)

        # Reuse existing sheet if present, else create a new one
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(title=sheet_name)
        # Ensure writer knows about the sheet
        writer.sheets[sheet_name] = worksheet

        # Write CM and tempo_licenca in the requested cells
        worksheet["E3"] = self.data_migracao
        worksheet["C5"] = self.funcionario.cm
        worksheet["C6"] = self.dados_faltantes_aeros.nome
        worksheet["C7"] = self.funcionario.data_admissao
        worksheet["D8"] = self.dados_faltantes_aeros.qtde_dias_licenca

        # Escreve progressões
        start_row = 13
        for i, prog in enumerate(self.progressoes):
            row = start_row + i
            worksheet.cell(row=row, column=2, value=prog.data)
            worksheet.cell(row=row, column=3, value=getattr(prog.nivel, "numero", None))
            worksheet.cell(row=row, column=4, value=prog.nota_avaliacao)
            worksheet.cell(row=row, column=5, value=prog.especial_concedida)
            worksheet.cell(row=row, column=6, value=prog.nota_media_especial)

        # Escreve resultado do reenquadramento
        if self.funcionario.dados_folha.classe == Classe.E1:
            worksheet["C34"] = "E2"
        else:
            worksheet["C34"] = self.funcionario.dados_folha.classe.name
        worksheet["C35"] = self.progressoes[-1].nivel.numero
        worksheet["C36"] = self._calcula_letra_maxima()
        worksheet["E37"] = self.progressoes[-1].data
        worksheet["E38"] = self.progressoes[-1].especial_concedida

    def _calcula_letra_maxima(self) -> str:
        """Compara as progs horizontais que servidor tem com o limite da sua carreira."""
        nivel = Nivel(
            self.progressoes[-1].nivel.numero,
            self.dados_faltantes_aeros.letra_maxima,
        )
        while True:
            try:
                self.funcionario.carreira.checa_nivel_valido(nivel=nivel)
                break
            except ValueError:
                nivel = nivel.anterior(0, 1)
        return nivel.letra


class TrajetoriasSimuladas:
    def __init__(
        self,
        funcionarios: Funcionario,
        data_migracao: date,
        avaliacoes: Avaliacoes = Avaliacoes.from_excel(),
    ):
        self.funcionarios = funcionarios  # {cm: Funcionario}
        self.data_migracao = data_migracao
        self.avaliacoes = avaliacoes  # Avaliacoes
        self.dados_faltantes = obtem_dados_faltantes_aeros()

        # Outputs
        self.trajetorias_simuladas = {}  # {cm: TrajetoriaSimulada}

    @classmethod
    def from_excel(cls, caminho_excel: str, data_migracao: date):
        """Cria uma instância de CMBH a partir de um arquivo Excel."""

        cmbh: CMBH = ImportadorProjecaoExcel().importa(
            caminho_excel, importa_folhas=False
        )
        return cls(cmbh.funcionarios, data_migracao)

    def calcula(self) -> None:
        """Calcula as carreiras simuladas de todos os funcionários."""
        for cm, funcionario in self.funcionarios.items():
            trajetoria_simulada = CalculaReenquadramento(
                funcionario,
                self.data_migracao,
                self.avaliacoes.do_cm(cm),
                self.dados_faltantes.get(cm, DadosFaltantesAeros("", 0, "")),
            ).calcula()

            self.trajetorias_simuladas[cm] = trajetoria_simulada

    def exporta_para_excel(self, caminho_excel: str) -> None:
        """Escreve todas as carreiras simuladas em um arquivo Excel.

        Abre um pandas.ExcelWriter para o caminho fornecido e escreve cada
        `TrajetoriaSimulada` em uma planilha separada (nome = CM). As planilhas
        são escritas em ordem crescente de CM.
        """
        # Garante que o diretório exista
        dirpath = os.path.dirname(caminho_excel)
        if dirpath:
            os.makedirs(dirpath, exist_ok=True)

        # Procura pelo arquivo de modelo
        possible_template = os.path.join(os.path.dirname(__file__), "modelo.xlsx")
        if not os.path.exists(possible_template):
            raise FileNotFoundError(f"Modelo não encontrado: {possible_template}")

        # Abre planilha de modelo
        template_wb = openpyxl.load_workbook(possible_template)
        template_ws = template_wb[template_wb.sheetnames[0]]

        for cm in sorted(self.trajetorias_simuladas.keys()):
            carreira_simulada: TrajetoriaSimulada = self.trajetorias_simuladas[cm]

            # Copia planilha de modelo
            new_ws = template_wb.copy_worksheet(template_ws)
            new_ws.title = str(cm)

            # Cria o writer
            writer = SimpleNamespace(book=template_wb, sheets={})
            writer.sheets[new_ws.title] = new_ws

            carreira_simulada.escreve(writer)

        # Salva uma cópia do workbook
        template_wb.save(caminho_excel)


class CalculaReenquadramento:
    def __init__(
        self,
        funcionario: Funcionario,
        data_migracao: date,
        avaliacoes: list[Avaliacao],
        dados_faltantes_aeros: DadosFaltantesAeros,
    ):
        self.funcionario = funcionario
        self.data_migracao = data_migracao
        self.avaliacoes = avaliacoes or []
        self.dados_faltantes_aeros = dados_faltantes_aeros
        self.indx_avaliacao = 0
        self.num_progressoes = 0

    def calcula(self) -> TrajetoriaSimulada:
        """Calcula a carreira simulada de um funcionário."""
        trajetoria_simulada = TrajetoriaSimulada(
            self.funcionario,
            self.data_migracao,
            [],
            self.dados_faltantes_aeros,
        )

        data_condicao_aposentadoria = (
            self.funcionario.aposentadoria.data_condicao_aposentadoria
        )
        progressao = self.funcionario.progressoes[-1]

        while progressao and progressao.data < self.data_migracao:

            prog_sim = self._converte_em_simulada(progressao)

            proxima_prog = self.funcionario.carreira.progride_verticalmente(progressao)
            nota = self._obtem_nota(proxima_prog)
            qtde_progressoes = self._qtde_progressoes(nota)

            # Verifica progressão especial
            especial_concedida, nota_media = self._processa_progressao_especial(nota)
            qtde_progressoes += int(especial_concedida)

            prog_sim.especial_concedida = especial_concedida
            prog_sim.nota_media_especial = nota_media
            prog_sim.nota_avaliacao = nota

            trajetoria_simulada.progressoes.append(prog_sim)

            # Avança a progressão de acordo com qtde_progressoes calculada
            progressao = self.funcionario.carreira.progride_verticalmente_por_quantidade_de_niveis(
                progressao, qtde_progressoes, data_condicao_aposentadoria
            )
            self.num_progressoes += 1

        self._remove_nota_ultima_avaliacao(trajetoria_simulada)
        return trajetoria_simulada

    def _obtem_nota(self, progressao: Progressao) -> int:
        """Retorna a nota a ser usada para a progressão, consumindo avaliações quando aplicável."""
        if not progressao:
            return 70
        if self._usa_nota_avaliacao(progressao.data):
            # Usa a avaliação disponível ou 70 como padrão
            nota = self._obtem_proxima_avaliacao_ou_padrao()
            return nota
        return 70

    def _processa_progressao_especial(self, nota: float) -> tuple[bool, float]:
        """Verifica se a progressão especial deve ser concedida e calcula a média."""
        if self._essa_progressao_pode_ter_especial():
            nota_anterior = self._nota_anterior()
            media = (nota + nota_anterior) / 2
            if media >= 70:
                return True, media
            print(
                f"CM {self.funcionario.cm}: Progressão especial não concedida, média {media} < 70"
            )
            return False, media
        return False, None

    def _obtem_proxima_avaliacao_ou_padrao(self) -> int:
        if self.avaliacoes and self.indx_avaliacao < len(self.avaliacoes):
            nota = self.avaliacoes[self.indx_avaliacao].nota
            self.indx_avaliacao += 1
            return nota
        self.indx_avaliacao += 1
        return 70

    def _essa_progressao_pode_ter_especial(self) -> bool:
        """Decide se deve avaliar possibilidade de progressão especial (quando índice é par)."""
        return self.num_progressoes % 2 == 1

    def _nota_anterior(self) -> int:
        """Retorna a nota da avaliação anterior quando existir, senão 70."""
        if self.indx_avaliacao <= 1:
            return 70
        prev_idx = self.indx_avaliacao - 2  # índice da avaliação anterior
        if self.avaliacoes and prev_idx < len(self.avaliacoes):
            return self.avaliacoes[prev_idx].nota
        return 70

    @staticmethod
    def _converte_em_simulada(progressao: Progressao) -> ProgressaoSimulada:
        """Converte uma Progressao em ProgressaoSimulada."""
        return ProgressaoSimulada(
            data=progressao.data,
            nivel=progressao.nivel,
            especial_concedida=False,
            nota_media_especial=None,
            nota_avaliacao=None,
        )

    @staticmethod
    def _remove_nota_ultima_avaliacao(trajetoria_simulada: TrajetoriaSimulada) -> None:
        """Remove a nota da última avaliação se a progressão não ocorreu."""
        if not trajetoria_simulada.progressoes:
            return
        last = trajetoria_simulada.progressoes[-1]
        last.nota_avaliacao = None
        last.nota_media_especial = None

    @staticmethod
    def _usa_nota_avaliacao(intersticio_fim: date) -> bool:
        """Se intersticio termina antes de 02 de abril de 2004, usa nota 70, caso
        contrário usa nota da avaliação.."""
        data_corte = date(2004, 4, 2)
        return intersticio_fim >= data_corte

    def _qtde_progressoes(self, nota: int) -> int:
        """Retorna o número de progressões em um reenquadramento."""
        if nota < 60:
            print(
                f"CM {self.funcionario.cm}: Progressão não concedida: nota {nota} < 60"
            )
            return 0
        if nota < 70:
            print(f"CM {self.funcionario.cm}: Somente uma progressão: nota {nota} < 70")
            return 1
        return 2
