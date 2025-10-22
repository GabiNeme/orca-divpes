import pandas as pd
from openpyxl.styles import Alignment, NamedStyle


class ExportadorExcel:
    """Classe para exportar DataFrames para Excel com formatação personalizada."""

    def __init__(self):
        self._configurar_estilo_numerico()

    def _configurar_estilo_numerico(self):
        """Configura o estilo de formatação para números."""
        self.estilo_numerico = NamedStyle(name="formato_numerico")
        self.estilo_numerico.number_format = "#,##0.00"

    def para_excel(
        self,
        df: pd.DataFrame,
        writer: pd.ExcelWriter,
        sheet_name: str,
        index: bool = False,
        **kwargs
    ) -> None:
        """
        Exporta DataFrame para Excel com formatação automática de números.

        Args:
            df: DataFrame a ser exportado
            writer: ExcelWriter do pandas
            sheet_name: Nome da planilha
            index: Se deve incluir o índice
            **kwargs: Argumentos adicionais para to_excel
        """
        # Primeiro, escreve o DataFrame normalmente
        df.to_excel(writer, sheet_name=sheet_name, index=index, **kwargs)

        # Depois aplica a formatação
        self._formatar_planilha(writer, sheet_name, df, index, **kwargs)

    def _formatar_planilha(
        self,
        writer: pd.ExcelWriter,
        sheet_name: str,
        df: pd.DataFrame,
        index: bool,
        **kwargs
    ) -> None:
        """Aplica formatação de números, quebra de linha automática e ajuste de colunas."""
        workbook = writer.book
        planilha = writer.sheets[sheet_name]

        # Adiciona o estilo ao workbook se não existir
        if "formato_numerico" not in workbook.named_styles:
            workbook.add_named_style(self.estilo_numerico)

        deslocamento_linha = kwargs.get("startrow", 0)
        deslocamento_coluna = kwargs.get("startcol", 0)
        primeira_coluna = (2 if index else 1) + deslocamento_coluna
        primeira_linha = 2 + deslocamento_linha

        # Aplica formatação para números float
        for idx_linha in range(len(df)):
            for idx_coluna, _ in enumerate(df.columns):
                celula = planilha.cell(
                    row=primeira_linha + idx_linha, column=primeira_coluna + idx_coluna
                )

                # Verifica se o valor é um float
                if self._eh_float(df.iloc[idx_linha, idx_coluna]):
                    celula.style = "formato_numerico"

        # Aplica wrap text na primeira linha (cabeçalhos)
        self._aplicar_quebra_linha_cabecalho(
            planilha, df, index, deslocamento_linha, deslocamento_coluna
        )

        # Aplica auto-fit na largura das colunas
        self._auto_ajustar_largura_colunas(planilha)

        # Aplica auto-fit na altura das linhas
        self._auto_ajustar_altura_linhas(planilha)

    def _aplicar_quebra_linha_cabecalho(
        self,
        worksheet,
        df: pd.DataFrame,
        index: bool,
        deslocamento_linha: int,
        deslocamento_coluna: int,
    ) -> None:
        """Aplica wrap text automático nos cabeçalhos."""
        linha_cabecalho = 1 + deslocamento_linha
        primeira_coluna = (2 if index else 1) + deslocamento_coluna

        # Aplica quebra de linha nos cabeçalhos das colunas
        for idx_coluna in range(len(df.columns)):
            celula = worksheet.cell(
                row=linha_cabecalho, column=primeira_coluna + idx_coluna
            )
            celula.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )

        if index:
            celula_indice = worksheet.cell(
                row=linha_cabecalho, column=1 + deslocamento_coluna
            )
            celula_indice.alignment = Alignment(
                wrap_text=True, vertical="center", horizontal="center"
            )

    def _auto_ajustar_largura_colunas(self, worksheet) -> None:
        """Ajusta a largura das colunas com base no conteúdo."""
        for coluna in worksheet.columns:
            comprimento_max = 0
            letra_coluna = coluna[0].column_letter
            for celula in coluna:
                if celula.value:
                    comprimento_celula = max(
                        len(str(linha)) for linha in str(celula.value).split("\n")
                    )
                    comprimento_max = max(comprimento_max, comprimento_celula)
            largura_ajustada = min(comprimento_max + 2, 50)
            worksheet.column_dimensions[letra_coluna].width = max(largura_ajustada, 10)

    def _auto_ajustar_altura_linhas(self, worksheet) -> None:
        """Ajusta a altura das linhas com base no conteúdo."""
        for linha in worksheet.iter_rows():
            max_linhas = 1
            for celula in linha:
                if celula.value and celula.alignment and celula.alignment.wrap_text:
                    linhas_necessarias = self._calcular_linhas_necessarias(
                        celula, worksheet
                    )
                    max_linhas = max(max_linhas, linhas_necessarias)
            if max_linhas > 1:
                worksheet.row_dimensions[linha[0].row].height = max_linhas * 15

    def _calcular_linhas_necessarias(self, celula, worksheet) -> int:
        """Calcula o número de linhas necessárias para exibir o conteúdo de uma célula."""
        if not celula.value:
            return 1
        texto = str(celula.value)
        letra_coluna = celula.column_letter
        largura_coluna = worksheet.column_dimensions[letra_coluna].width or 10
        linhas_explicitas = texto.count("\n") + 1
        caracteres_por_linha = max(int(largura_coluna), 1)
        linhas_wrap_automatico = 0
        for linha in texto.split("\n"):
            comprimento_linha = len(linha)
            linhas_wrap_automatico += max(
                1,
                (comprimento_linha + caracteres_por_linha - 1) // caracteres_por_linha,
            )
        return max(linhas_explicitas, linhas_wrap_automatico)

    def _eh_float(self, valor) -> bool:
        """Verifica se um valor é um número float."""
        import numpy as np

        # Verifica se é None ou NaN
        if valor is None or pd.isna(valor):
            return False

        # Verifica se é um tipo float (exclui int, bool, etc.)
        return isinstance(valor, (float, np.floating))


# Instância global para uso conveniente
exportador_excel = ExportadorExcel()


# Função de conveniência
def para_excel_formatado(
    df: pd.DataFrame,
    writer: pd.ExcelWriter,
    sheet_name: str,
    index: bool = False,
    **kwargs
) -> None:
    """
    Função de conveniência para exportar DataFrame com formatação.

    Args:
        df: DataFrame a ser exportado
        writer: ExcelWriter do pandas
        sheet_name: Nome da planilha
        index: Se deve incluir o índice
        **kwargs: Argumentos adicionais para to_excel
    """
    exportador_excel.para_excel(df, writer, sheet_name, index, **kwargs)
